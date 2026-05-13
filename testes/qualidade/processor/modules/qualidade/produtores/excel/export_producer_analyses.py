from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any


OPERATION = "qualidade.exportar_produtores_analises"

HEADERS = [
    "Código",
    "Produtor",
    "Cidade",
    "Rota",
    "Status",
    "Data da análise",
    "CCS",
    "UFC",
    "Gordura",
    "Proteína",
    "Lactose",
    "Sólidos totais",
    "Caseína",
    "SNG",
    "Ureia",
    "ATB",
    "BCL",
    "Temperatura",
]

METRIC_KEYS = [
    "ccs",
    "ufc",
    "gordura",
    "proteina",
    "lactose",
    "solidos_totais",
    "caseina",
    "sng",
    "ureia",
    "antibiotico",
    "bacteria",
    "temperatura",
]


def load_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("Dependência openpyxl não instalada para exportação XLSX.") from exc

    return {
        "Workbook": Workbook,
        "Image": Image,
        "Alignment": Alignment,
        "Border": Border,
        "Font": Font,
        "PatternFill": PatternFill,
        "Side": Side,
        "get_column_letter": get_column_letter,
    }


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def date_br(value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    try:
        return datetime.strptime(text[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        return text


def decimal_value(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def normalized_metric(key: str, value: Any) -> Any:
    if key in {"antibiotico", "bacteria"}:
        if value is None or value == "":
            return ""
        try:
            return "POS" if float(value) > 0 else "NEG"
        except (TypeError, ValueError):
            return str(value)

    number = decimal_value(value)
    if number is None:
        return None
    if key in {"ccs", "ufc"}:
        return int(number)
    return number


def producer_row(produtor: dict[str, Any]) -> list[Any]:
    analysis = produtor.get("ultima_analise") or {}
    status = "Com análise" if analysis else "Sem análise"

    row = [
        normalize_text(produtor.get("codigo")),
        normalize_text(produtor.get("nome")),
        normalize_text(produtor.get("cidade")),
        normalize_text(produtor.get("rota")),
        status,
        date_br(analysis.get("data")),
    ]
    row.extend(normalized_metric(key, analysis.get(key)) for key in METRIC_KEYS)
    return row


def apply_standard_header(ws, deps: dict[str, Any], payload: dict[str, Any], title: str, logo_path: Path | None) -> None:
    Alignment = deps["Alignment"]
    Font = deps["Font"]
    PatternFill = deps["PatternFill"]

    ws.sheet_view.showGridLines = False
    ws.merge_cells("D1:R1")
    ws.merge_cells("D2:R2")
    ws.merge_cells("D3:R3")
    ws["D1"] = title
    ws["D2"] = f"Período de referência: {payload['periodo']['label']}"
    ws["D3"] = f"Gerado em: {date_br(payload.get('gerado_em', ''))} {normalize_text(payload.get('gerado_hora'))}"

    ws["D1"].font = Font(name="Arial", size=15, bold=True, color="1F2937")
    ws["D2"].font = Font(name="Arial", size=10, color="6B7280")
    ws["D3"].font = Font(name="Arial", size=10, color="6B7280")
    for row in range(1, 5):
        ws[f"D{row}"].alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 18

    if logo_path and logo_path.exists():
        Image = deps["Image"]
        logo = Image(str(logo_path))
        logo.width = 150
        logo.height = 58
        ws.add_image(logo, "B2")

    for cell in ws[5]:
        cell.fill = PatternFill("solid", fgColor="111827")


def style_table(ws, deps: dict[str, Any], table_ref: str, table_name: str, header_row: int = 5) -> None:
    Alignment = deps["Alignment"]
    Border = deps["Border"]
    Font = deps["Font"]
    PatternFill = deps["PatternFill"]
    Side = deps["Side"]

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=9, color="111827")
    border = Border(bottom=Side(style="thin", color="E5E7EB"))
    stripe_fill = PatternFill("solid", fgColor="F9FAFB")

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
        for cell in row:
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if row_index % 2 == 0:
                cell.fill = stripe_fill

    ws.freeze_panes = f"A{header_row + 1}"


def apply_widths(ws) -> None:
    widths = [10, 34, 22, 12, 15, 16, 10, 10, 11, 11, 11, 14, 11, 10, 10, 8, 8, 13]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=5, column=index).column_letter].width = width


def write_table_rows(ws, headers: list[str], rows: list[list[Any]], start_row: int = 5) -> int:
    for column_index, header in enumerate(headers, start=1):
        ws.cell(row=start_row, column=column_index, value=header)

    for row_index, row in enumerate(rows, start=start_row + 1):
        for column_index, value in enumerate(row, start=1):
            ws.cell(row=row_index, column=column_index, value=value)

    return max(start_row + len(rows), start_row + 1)


def build_workbook(payload: dict[str, Any], output_path: Path, logo_path: Path | None) -> dict[str, Any]:
    deps = load_openpyxl()
    Workbook = deps["Workbook"]
    get_column_letter = deps["get_column_letter"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Produtores e análises"
    apply_standard_header(ws, deps, payload, "Relatório de produtores e análises", logo_path)

    produtores = payload.get("produtores", [])
    rows = [producer_row(produtor) for produtor in produtores]
    last_row = write_table_rows(ws, HEADERS, rows)
    last_col = get_column_letter(len(HEADERS))
    style_table(ws, deps, f"A5:{last_col}{last_row}", "TabelaProdutoresAnalises")
    apply_widths(ws)

    ws_summary = wb.create_sheet("Resumo")
    apply_standard_header(ws_summary, deps, payload, "Resumo da exportação", logo_path)
    summary_rows = [
        ("Produtores exportados", payload["totais"]["produtores"]),
        ("Com análise no mês", payload["totais"]["com_analise"]),
        ("Sem análise no mês", payload["totais"]["sem_analise"]),
        ("Mês de referência", payload["periodo"]["label"]),
    ]
    summary_last_row = write_table_rows(ws_summary, ["Indicador", "Valor"], [list(row) for row in summary_rows])
    style_table(ws_summary, deps, f"A5:B{summary_last_row}", "TabelaResumoExportacao")
    ws_summary.column_dimensions["A"].width = 32
    ws_summary.column_dimensions["B"].width = 22

    sem_analise = [row for row in rows if row[4] == "Sem análise"]
    ws_missing = wb.create_sheet("Sem análise")
    apply_standard_header(ws_missing, deps, payload, "Produtores sem análise no período", logo_path)
    missing_headers = ["Código", "Produtor", "Cidade", "Rota", "Status"]
    missing_last_row = write_table_rows(ws_missing, missing_headers, [row[:5] for row in sem_analise])
    style_table(ws_missing, deps, f"A5:E{missing_last_row}", "TabelaSemAnalise")
    for column, width in zip(["A", "B", "C", "D", "E"], [10, 34, 22, 12, 15], strict=True):
        ws_missing.column_dimensions[column].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return {
        "success": True,
        "operation": OPERATION,
        "output": str(output_path),
        "summary": payload["totais"],
    }


def export_file(input_path: Path, output_path: Path, logo_path: Path | None) -> dict[str, Any]:
    try:
        payload = json.loads(input_path.read_text(encoding="utf-8"))
        return build_workbook(payload, output_path, logo_path)
    except Exception as exc:
        return {
            "success": False,
            "operation": OPERATION,
            "output": str(output_path),
            "errors": [
                {
                    "code": "EXPORT_810",
                    "message": "Falha ao gerar planilha.",
                    "details": {"error": str(exc)},
                }
            ],
        }


def main() -> int:
    parser = argparse.ArgumentParser(description="Exporta produtores e análises do Santi'Lac.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--logo")
    args = parser.parse_args()

    result = export_file(
        Path(args.input),
        Path(args.output),
        Path(args.logo) if args.logo else None,
    )
    print(json.dumps(result, ensure_ascii=False))
    return 0 if result["success"] else 1


if __name__ == "__main__":
    sys.exit(main())
