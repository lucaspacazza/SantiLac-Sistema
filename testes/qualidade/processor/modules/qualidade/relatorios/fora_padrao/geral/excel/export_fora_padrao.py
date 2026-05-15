from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any


OPERATION = "qualidade.relatorios.fora_padrao.geral.excel"

HEADERS = [
    "Indicador",
    "Código",
    "Produtor",
    "Cidade",
    "Rota",
    "Data da análise",
    "Valor",
    "Referência",
    "Unidade",
    "Gravidade",
]


def load_openpyxl() -> dict[str, Any]:
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError("Dependência openpyxl não instalada para exportação XLSX.") from exc

    return {
        "Workbook": Workbook,
        "Image": Image,
        "Alignment": Alignment,
        "Border": Border,
        "Font": Font,
        "PatternFill": PatternFill,
        "Side": Side,
    }


def text(value: Any) -> str:
    return "" if value is None else str(value)


def date_br(value: Any) -> str:
    raw = text(value)
    if not raw:
        return ""
    try:
        return datetime.strptime(raw[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        return raw


def number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def apply_header(ws, deps: dict[str, Any], payload: dict[str, Any], logo_path: Path | None) -> None:
    Alignment = deps["Alignment"]
    Font = deps["Font"]

    ws.sheet_view.showGridLines = False
    ws.merge_cells("D1:J1")
    ws.merge_cells("D2:J2")
    ws.merge_cells("D3:J3")
    ws["D1"] = text(payload.get("titulo")) or "Relatório de produtores fora do padrão"
    ws["D2"] = f"Período de referência: {payload['periodo']['label']}"
    ws["D3"] = f"Gerado em: {date_br(payload.get('gerado_em'))} {text(payload.get('gerado_hora'))}"

    ws["D1"].font = Font(name="Arial", size=15, bold=True, color="111827")
    ws["D2"].font = Font(name="Arial", size=10, color="6B7280")
    ws["D3"].font = Font(name="Arial", size=10, color="6B7280")
    for row in range(1, 4):
        ws[f"D{row}"].alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 18

    if logo_path and logo_path.exists():
        Image = deps["Image"]
        logo = Image(str(logo_path))
        logo.width = 150
        logo.height = 58
        ws.add_image(logo, "B2")


def style_sheet(ws, deps: dict[str, Any], header_row: int = 5) -> None:
    Alignment = deps["Alignment"]
    Border = deps["Border"]
    Font = deps["Font"]
    PatternFill = deps["PatternFill"]
    Side = deps["Side"]

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=9, color="111827")
    border = Border(bottom=Side(style="thin", color="E5E7EB"))
    stripe_fill = PatternFill("solid", fgColor="F9FAFB")

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
        for cell in row:
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if row_index % 2 == 0:
                cell.fill = stripe_fill

    ws.freeze_panes = f"A{header_row + 1}"


def row_from_item(item: dict[str, Any]) -> list[Any]:
    return [
        text(item.get("indicador_label")),
        text(item.get("codigo")),
        text(item.get("nome")),
        text(item.get("cidade")),
        text(item.get("rota")),
        date_br(item.get("data")),
        number(item.get("valor")),
        text(item.get("referencia")),
        text(item.get("unidade")),
        number(item.get("gravidade")),
    ]


def build_workbook(payload: dict[str, Any], output_path: Path, logo_path: Path | None) -> dict[str, Any]:
    deps = load_openpyxl()
    Workbook = deps["Workbook"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Fora do padrão"
    apply_header(ws, deps, payload, logo_path)

    for column, header in enumerate(HEADERS, start=1):
        ws.cell(row=5, column=column, value=header)

    items = payload.get("items", [])
    for row_index, item in enumerate(items, start=6):
        for column, value in enumerate(row_from_item(item), start=1):
            ws.cell(row=row_index, column=column, value=value)

    widths = [24, 10, 34, 22, 10, 16, 12, 18, 10, 12]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=5, column=index).column_letter].width = width

    style_sheet(ws, deps)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return {
        "success": True,
        "operation": OPERATION,
        "output": str(output_path),
        "summary": payload.get("totais", {}),
    }


def export_file(input_path: Path, output_path: Path, logo_path: Path | None) -> dict[str, Any]:
    try:
        payload = json.loads(input_path.read_text(encoding="utf-8"))
        return build_workbook(payload, output_path, logo_path)
    except Exception as exc:
        return {
            "success": False,
            "operation": OPERATION,
            "output": str(output_path),
            "errors": [{"code": "EXPORT_910", "message": "Falha ao gerar planilha.", "details": {"error": str(exc)}}],
        }


def main() -> int:
    parser = argparse.ArgumentParser(description="Exporta relatório geral de produtores fora do padrão.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--logo")
    args = parser.parse_args()

    result = export_file(Path(args.input), Path(args.output), Path(args.logo) if args.logo else None)
    print(json.dumps(result, ensure_ascii=False))
    return 0 if result["success"] else 1


if __name__ == "__main__":
    sys.exit(main())
