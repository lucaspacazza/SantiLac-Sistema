from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.graphics.barcode import qr
from reportlab.graphics.shapes import Drawing
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


HEADERS = [
    ("Ordem", "ordem"),
    ("Cliente", "cliente"),
    ("Destino", "destino"),
    ("Data prevista", "data_prevista"),
    ("Status", "status"),
    ("Palete", "palete"),
    ("Produto", "produto"),
    ("Caixas", "caixas"),
    ("Peso (kg)", "peso_total"),
    ("QR Code", "qr_code"),
    ("Conferência", "conferencia"),
    ("Operador", "operador_conferencia"),
]


def load_items(path: Path) -> list[dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    items = payload.get("itens", [])
    return items if isinstance(items, list) else []


def export_xlsx(items: list[dict[str, Any]], output: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Expedição"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(2, len(items) + 1)}"

    for column, (label, _) in enumerate(HEADERS, 1):
        cell = sheet.cell(1, column, label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(vertical="center")

    for row_index, item in enumerate(items, 2):
        for column, (_, key) in enumerate(HEADERS, 1):
            value = item.get(key, "")
            cell = sheet.cell(row_index, column, value)
            cell.alignment = Alignment(vertical="top")
            if key == "peso_total":
                cell.number_format = '#,##0.000'

    widths = [18, 24, 28, 14, 14, 10, 22, 10, 14, 36, 16, 22]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    workbook.save(output)


def qr_drawing(value: str, size: float = 22 * mm) -> Drawing:
    widget = qr.QrCodeWidget(value or "-")
    bounds = widget.getBounds()
    width = bounds[2] - bounds[0]
    height = bounds[3] - bounds[1]
    drawing = Drawing(size, size, transform=[size / width, 0, 0, size / height, 0, 0])
    drawing.add(widget)
    return drawing


def export_pdf(items: list[dict[str, Any]], output: Path, logo: Path | None) -> None:
    document = SimpleDocTemplate(
        str(output),
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title="Relatório de Expedição",
    )
    styles = getSampleStyleSheet()
    story: list[Any] = []

    header: list[Any] = []
    if logo and logo.exists():
        header.append(Image(str(logo), width=34 * mm, height=14 * mm))
    header.append(Paragraph("<b>Relatório de Expedição</b><br/><font size='8'>Gerado em "
                            + datetime.now().strftime("%d/%m/%Y %H:%M") + "</font>", styles["Normal"]))
    story.append(Table([header], colWidths=[42 * mm, 220 * mm] if len(header) == 2 else [262 * mm]))
    story.append(Spacer(1, 5 * mm))

    columns = ["Ordem", "Cliente", "Destino", "Status", "Palete", "Produto", "Caixas", "Peso"]
    table_data = [columns]
    for item in items:
        table_data.append([
            item.get("ordem", ""),
            item.get("cliente", ""),
            item.get("destino", ""),
            item.get("status", ""),
            item.get("palete", ""),
            item.get("produto", ""),
            item.get("caixas", 0),
            f"{float(item.get('peso_total', 0) or 0):,.3f}".replace(",", "X").replace(".", ",").replace("X", "."),
        ])

    table = Table(table_data, repeatRows=1, colWidths=[26*mm, 39*mm, 48*mm, 23*mm, 17*mm, 37*mm, 15*mm, 22*mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d1d5db")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f4f6")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(table)

    qr_items = [item for item in items if item.get("qr_code")]
    if qr_items:
        story.append(PageBreak())
        story.append(Paragraph("<b>QR Codes dos paletes</b>", styles["Heading2"]))
        story.append(Spacer(1, 3 * mm))
        cards = []
        row = []
        for item in qr_items:
            label = Paragraph(
                f"<b>Palete {item.get('palete', '')}</b><br/>{item.get('produto', '')}<br/>"
                f"{item.get('ordem', '')} | {float(item.get('peso_total', 0) or 0):.3f} kg",
                styles["Normal"],
            )
            row.append(Table([[qr_drawing(str(item["qr_code"])), label]], colWidths=[25*mm, 54*mm]))
            if len(row) == 3:
                cards.append(row)
                row = []
        if row:
            row.extend([""] * (3 - len(row)))
            cards.append(row)
        story.append(Table(cards, colWidths=[86*mm, 86*mm, 86*mm], style=[
            ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("PADDING", (0, 0), (-1, -1), 4),
        ]))

    document.build(story)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--logo")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    items = load_items(input_path)
    if output_path.suffix.lower() == ".pdf":
        export_pdf(items, output_path, Path(args.logo) if args.logo else None)
    else:
        export_xlsx(items, output_path)

    print(json.dumps({
        "success": True,
        "arquivo": output_path.name,
        "linhas": len(items),
    }, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
