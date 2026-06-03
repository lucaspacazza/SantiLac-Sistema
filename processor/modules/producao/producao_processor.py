#!/usr/bin/env python3
import argparse
import json
import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

OOXML_NAMESPACES = {
    "wpc": "http://schemas.microsoft.com/office/word/2010/wordprocessingCanvas",
    "cx": "http://schemas.microsoft.com/office/drawing/2014/chartex",
    "cx1": "http://schemas.microsoft.com/office/drawing/2015/9/8/chartex",
    "cx2": "http://schemas.microsoft.com/office/drawing/2015/10/21/chartex",
    "cx3": "http://schemas.microsoft.com/office/drawing/2016/5/9/chartex",
    "cx4": "http://schemas.microsoft.com/office/drawing/2016/5/10/chartex",
    "cx5": "http://schemas.microsoft.com/office/drawing/2016/5/11/chartex",
    "cx6": "http://schemas.microsoft.com/office/drawing/2016/5/12/chartex",
    "cx7": "http://schemas.microsoft.com/office/drawing/2016/5/13/chartex",
    "cx8": "http://schemas.microsoft.com/office/drawing/2016/5/14/chartex",
    "mc": "http://schemas.openxmlformats.org/markup-compatibility/2006",
    "aink": "http://schemas.microsoft.com/office/drawing/2016/ink",
    "am3d": "http://schemas.microsoft.com/office/drawing/2017/model3d",
    "o": "urn:schemas-microsoft-com:office:office",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "m": "http://schemas.openxmlformats.org/officeDocument/2006/math",
    "v": "urn:schemas-microsoft-com:vml",
    "wp14": "http://schemas.microsoft.com/office/word/2010/wordprocessingDrawing",
    "wp": "http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing",
    "w10": "urn:schemas-microsoft-com:office:word",
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
    "w14": "http://schemas.microsoft.com/office/word/2010/wordml",
    "w15": "http://schemas.microsoft.com/office/word/2012/wordml",
    "w16cex": "http://schemas.microsoft.com/office/word/2018/wordml/cex",
    "w16cid": "http://schemas.microsoft.com/office/word/2016/wordml/cid",
    "w16": "http://schemas.microsoft.com/office/word/2018/wordml",
    "w16se": "http://schemas.microsoft.com/office/word/2015/wordml/symex",
    "wpg": "http://schemas.microsoft.com/office/word/2010/wordprocessingGroup",
    "wpi": "http://schemas.microsoft.com/office/word/2010/wordprocessingInk",
    "wne": "http://schemas.microsoft.com/office/word/2006/wordml",
    "wps": "http://schemas.microsoft.com/office/word/2010/wordprocessingShape",
}

NS = {"w": OOXML_NAMESPACES["w"]}
for prefix, uri in OOXML_NAMESPACES.items():
    ET.register_namespace(prefix, uri)

BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_DIR = BASE_DIR / "templates"
LOGO_PATH = BASE_DIR / "assets" / "logo.png"
LOGO_TARGET = "word/media/santilac_logo.png"

FORM_CONFIG = {
    "formulacao_queijo": {
        "template": "plan_6_3_formulacao_queijo.docx",
        "documento": "PLAN_6.3",
        "slug": "formulacao-queijo",
    },
    "soro_refrigerado": {
        "template": "plan_6_7_soro_refrigerado.docx",
        "documento": "PLAN_6.7",
        "slug": "soro-refrigerado",
    },
    "formulacao_creme": {
        "template": "plan_6_9_formulacao_creme.docx",
        "documento": "PLAN_6.9",
        "slug": "formulacao-creme",
    },
    "producao_creme": {
        "template": "plan_6_10_producao_creme.docx",
        "documento": "PLAN_6.10",
        "slug": "producao-creme",
    },
}

OP_FIELDS = [
    "PRODUÇÃO DIARIA / DATA",
    "LTS PRODUZIDOS TOTAL",
    "LOTE DO QUEIJO",
    "PEÇAS F4",
    "PEÇAS F1",
    "PEÇAS F6",
    "PEÇAS COLONIAL",
    "PEÇAS COALHO",
    "PEÇAS PROVOLONE",
    "PEÇAS GOUDA",
    "PEÇAS GRUYERE",
    "PEÇAS PRATO",
    "CLORETO CÁLCIO",
    "COALHO",
    "FERMENTO (MVD)",
    "FERMENTO (FAST)",
    "FERMENTO",
    "FERMENTO",
    "CORANTE",
]


def w_tag(name):
    return f"{{{NS['w']}}}{name}"


def text_of(element):
    return "".join(node.text or "" for node in element.findall(".//w:t", NS)).strip()


def cells(row):
    return row.findall("./w:tc", NS)


def rows(table):
    return table.findall("./w:tr", NS)


def first_paragraph(cell):
    paragraph = cell.find("./w:p", NS)
    if paragraph is None:
        paragraph = ET.SubElement(cell, w_tag("p"))
    return paragraph


def clear_paragraph_text(paragraph):
    for child in list(paragraph):
        if child.tag != w_tag("pPr"):
            paragraph.remove(child)


def set_cell_text(cell, value):
    paragraph = first_paragraph(cell)
    for child in list(cell):
        if child.tag == w_tag("p") and child is not paragraph:
            cell.remove(child)
    clear_paragraph_text(paragraph)
    run = ET.SubElement(paragraph, w_tag("r"))
    text = ET.SubElement(run, w_tag("t"))
    text.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    text.text = value if value is not None else ""


def set_cell_lines(cell, lines):
    paragraph = first_paragraph(cell)
    for child in list(cell):
        if child.tag == w_tag("p") and child is not paragraph:
            cell.remove(child)
    clear_paragraph_text(paragraph)
    run = ET.SubElement(paragraph, w_tag("r"))
    for index, line in enumerate(lines):
        if index > 0:
            ET.SubElement(run, w_tag("br"))
        text = ET.SubElement(run, w_tag("t"))
        text.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
        text.text = line if line is not None else ""


def append_to_label(cell, value):
    label = text_of(cell)
    suffix = value if value is not None else ""
    set_cell_text(cell, f"{label} {suffix}".strip())


def clone_row(row):
    return ET.fromstring(ET.tostring(row, encoding="utf-8"))


def find_table(root, index):
    return root.findall(".//w:tbl", NS)[index]


def find_row_by_label(table, label):
    wanted = normalize(label)
    wanted_compact = compact_normalize(label)
    for row in rows(table):
        row_cells = cells(row)
        if row_cells and (
            normalize(text_of(row_cells[0])).startswith(wanted)
            or compact_normalize(text_of(row_cells[0])).startswith(wanted_compact)
        ):
            return row
    return None


def find_rows_by_label(table, label):
    wanted = normalize(label)
    wanted_compact = compact_normalize(label)
    found = []
    for row in rows(table):
        row_cells = cells(row)
        if row_cells and (
            normalize(text_of(row_cells[0])).startswith(wanted)
            or compact_normalize(text_of(row_cells[0])).startswith(wanted_compact)
        ):
            found.append(row)
    return found


def normalize(value):
    return re.sub(r"\s+", " ", value or "").strip().lower()


def compact_normalize(value):
    return re.sub(r"\s+", "", normalize(value))


def value(data, key, suffix=""):
    raw = data.get(key)
    if raw is None or raw == "":
        return ""
    if isinstance(raw, float):
        text = f"{raw:.3f}".rstrip("0").rstrip(".")
    else:
        text = str(raw)
    return f"{text}{suffix}"


def date_br(value):
    if not value:
        return ""
    parts = str(value).split("-")
    if len(parts) == 3:
        return f"{parts[2]}/{parts[1]}/{parts[0]}"
    return str(value)


def status_label(raw):
    labels = {
        "negativo": "Negativo",
        "positivo": "Positivo",
        "nao_aplicavel": "Não aplicável",
    }
    return labels.get(raw or "", raw or "")


def month_label(value):
    months = {
        1: "Janeiro",
        2: "Fevereiro",
        3: "Março",
        4: "Abril",
        5: "Maio",
        6: "Junho",
        7: "Julho",
        8: "Agosto",
        9: "Setembro",
        10: "Outubro",
        11: "Novembro",
        12: "Dezembro",
    }
    if value is None or value == "":
        return ""
    try:
        return months.get(int(value), str(value))
    except ValueError:
        return str(value)


def grouped_insumos(data):
    grouped = {}
    for item in data.get("insumos") or []:
        kind = item.get("tipo_insumo") or "outro"
        quantity = item.get("quantidade")
        unit = item.get("unidade") or ""
        lot = item.get("lote_insumo") or ""
        quantity_text = "" if quantity in (None, "") else f"{quantity:g} {unit}".strip()
        grouped.setdefault(kind, {"quantidades": [], "lotes": []})
        if quantity_text:
            grouped[kind]["quantidades"].append(quantity_text)
        if lot:
            grouped[kind]["lotes"].append(str(lot))
    return grouped


def fill_label_value_row(table, label, filled_value, cell_index=1):
    row = find_row_by_label(table, label)
    if row is None:
        return
    row_cells = cells(row)
    if len(row_cells) > cell_index:
        set_cell_text(row_cells[cell_index], filled_value)


def page_break_paragraph():
    paragraph = ET.Element(w_tag("p"))
    run = ET.SubElement(paragraph, w_tag("r"))
    br = ET.SubElement(run, w_tag("br"))
    br.set(w_tag("type"), "page")
    return paragraph


def is_empty_paragraph(element):
    if element.tag != w_tag("p"):
        return False
    if text_of(element):
        return False
    return not element.findall(".//w:br", NS) and not element.findall(".//w:drawing", NS)


def remove_trailing_empty_paragraphs(root):
    body = root.find(".//w:body", NS)
    if body is None:
        return
    children = list(body)
    while len(children) >= 2 and children[-1].tag == w_tag("sectPr") and is_empty_paragraph(children[-2]):
        body.remove(children[-2])
        children = list(body)


def insert_row_before(table, target_row, new_row):
    table.insert(list(table).index(target_row), new_row)


def ensure_gordura_inicial_row(table):
    gordura_inicial_row = find_row_by_label(table, "Gordura Inicial")
    gordura_final_row = find_row_by_label(table, "Gordura Final")
    if gordura_inicial_row is not None or gordura_final_row is None:
        return
    initial_row = clone_row(gordura_final_row)
    initial_cells = cells(initial_row)
    if initial_cells:
        set_cell_text(initial_cells[0], "Gordura Inicial")
    for cell in initial_cells[1:]:
        set_cell_text(cell, "")
    insert_row_before(table, gordura_final_row, initial_row)


def clear_value_columns(table):
    for row in rows(table):
        for cell in cells(row)[1:]:
            set_cell_text(cell, "")


def set_row_height(row, height):
    tr_pr = row.find("./w:trPr", NS)
    if tr_pr is None:
        tr_pr = ET.Element(w_tag("trPr"))
        row.insert(0, tr_pr)
    tr_height = tr_pr.find("./w:trHeight", NS)
    if tr_height is None:
        tr_height = ET.SubElement(tr_pr, w_tag("trHeight"))
    tr_height.set(w_tag("val"), str(height))
    tr_height.set(w_tag("hRule"), "exact")


def compact_formulacao_table(table):
    for index, row in enumerate(rows(table)):
        set_row_height(row, 520 if index == 0 else 390)


def ensure_ignorable_namespace_declarations(xml_bytes):
    xml = xml_bytes.decode("utf-8")
    match = re.search(r"\bmc:Ignorable=\"([^\"]+)\"", xml)
    if not match:
        return xml_bytes

    document_start = xml.find("<w:document")
    document_end = xml.find(">", document_start)
    if document_start == -1 or document_end == -1:
        return xml_bytes

    document_tag = xml[document_start:document_end]
    missing_declarations = []
    for prefix in match.group(1).split():
        if f"xmlns:{prefix}=" not in document_tag and prefix in OOXML_NAMESPACES:
            missing_declarations.append(f' xmlns:{prefix}="{OOXML_NAMESPACES[prefix]}"')

    if not missing_declarations:
        return xml_bytes

    xml = xml[:document_end] + "".join(missing_declarations) + xml[document_end:]
    return xml.encode("utf-8")


def fill_formulacao_queijo(root, data):
    if isinstance(data.get("items"), list):
        fill_formulacao_queijo_diario(root, data)
        return

    table = find_table(root, 0)
    table_rows = rows(table)
    if table_rows:
        first_cells = cells(table_rows[0])
        if first_cells:
            set_cell_text(first_cells[0], "Tipo de Queijo")
        if len(first_cells) > 1:
            set_cell_text(first_cells[1], data.get("tipo_queijo") or "")
        if len(first_cells) > 3:
            set_cell_text(first_cells[3], "Data:")
        if len(first_cells) > 4:
            set_cell_text(first_cells[4], date_br(data.get("data_formulacao")))

    fill_label_value_row(table, "Silo", data.get("silo") or "")
    fill_label_value_row(table, "Lote do Leite", data.get("lote_leite") or "")
    fill_label_value_row(table, "Lote do Queijo", data.get("lote_queijo") or "")
    fill_label_value_row(table, "N° Queijomatic", data.get("numero_queijomatic") or "")
    fill_label_value_row(table, "Início Enchimento", data.get("inicio_enchimento") or "")
    fill_label_value_row(table, "Quantidade Leite", value(data, "quantidade_leite", " L"))
    fill_label_value_row(table, "Temperatura. De Pasteurização", value(data, "temperatura_pasteurizacao", " °C"))
    fill_label_value_row(table, "Fosfatase", status_label(data.get("fosfatase")))
    fill_label_value_row(table, "Peroxidase", status_label(data.get("peroxidase")))

    insumos = grouped_insumos(data)
    mapping = {
        "fermento_mvd": ("Quantidade de Fermento (MVD)", "Lote do Fermento"),
        "fermento_fast": ("Quantidade de Fermento (FAST)", "Lote do Fermento"),
        "fermento": ("Quantidade de Fermento", "Lote do Fermento"),
        "cloreto": ("Quantidade de Cloreto", "Lote do Cloreto"),
        "corante": ("Quantidade de Corante", "Lote do Corante"),
        "coalho": ("Quantidade de Coalho", "Lote do Coalho"),
    }
    used_lote_rows = {}
    for kind, (quantity_label, lot_label) in mapping.items():
        item = insumos.get(kind, {})
        fill_label_value_row(table, quantity_label, "; ".join(item.get("quantidades", [])))
        lote_rows = [row for row in rows(table) if cells(row) and normalize(text_of(cells(row)[0])).startswith(normalize(lot_label))]
        target_index = used_lote_rows.get(lot_label, 0)
        if target_index < len(lote_rows):
            target_cells = cells(lote_rows[target_index])
            if len(target_cells) > 1:
                set_cell_text(target_cells[1], "; ".join(item.get("lotes", [])))
        used_lote_rows[lot_label] = target_index + 1

    gordura_final_row = find_row_by_label(table, "Gordura Final")
    gordura_inicial_row = find_row_by_label(table, "Gordura Inicial")
    if gordura_inicial_row is None and gordura_final_row is not None:
        initial_row = clone_row(gordura_final_row)
        initial_cells = cells(initial_row)
        if initial_cells:
            set_cell_text(initial_cells[0], "Gordura Inicial")
        if len(initial_cells) > 1:
            set_cell_text(initial_cells[1], value(data, "gordura_inicial", " %"))
        table.insert(list(table).index(gordura_final_row), initial_row)
    else:
        fill_label_value_row(table, "Gordura Inicial", value(data, "gordura_inicial", " %"))

    fill_label_value_row(table, "Gordura Final", value(data, "gordura_final", " %"))
    fill_label_value_row(table, "Acidez", value(data, "acidez", " °D"))
    fill_label_value_row(table, "Temperatura da Coagulação", value(data, "temperatura_coagulacao", " °C"))
    fill_label_value_row(table, "Hora da Coagulação", data.get("hora_coagulacao") or "")
    fill_label_value_row(table, "Hora do Corte", data.get("hora_corte") or "")
    fill_label_value_row(table, "Temperatura de Cozimento", value(data, "temperatura_cozimento", " °C"))
    fill_label_value_row(table, "Responsável pela Produção", data.get("responsavel") or "")


def fill_formulacao_queijo_diario(root, data):
    items = data.get("items") or []
    table = find_table(root, 0)
    body = root.find(".//w:body", NS)
    if body is None:
        return

    chunks = [items[index:index + 6] for index in range(0, len(items), 6)] or [[]]
    insert_at = list(body).index(table) + 1
    all_tables = [table]

    for _ in chunks[1:]:
        cloned_table = clone_row(table)
        body.insert(insert_at, page_break_paragraph())
        insert_at += 1
        body.insert(insert_at, cloned_table)
        insert_at += 1
        all_tables.append(cloned_table)

    for current_table, chunk in zip(all_tables, chunks):
        fill_formulacao_queijo_table(current_table, data.get("data_formulacao"), chunk)
    remove_trailing_empty_paragraphs(root)


def fill_formulacao_queijo_table(table, data_formulacao, items):
    ensure_gordura_inicial_row(table)
    clear_value_columns(table)
    compact_formulacao_table(table)

    table_rows = rows(table)
    if table_rows:
        first_cells = cells(table_rows[0])
        if first_cells:
            set_cell_lines(first_cells[0], ["Tipo de Queijo", f"Data: {date_br(data_formulacao)}"])
        for column, item in enumerate(items[:6], start=1):
            if column < len(first_cells):
                set_cell_text(first_cells[column], item.get("tipo_queijo") or "")

    simple_fields = [
        ("Silo", "silo", ""),
        ("Lote do Leite", "lote_leite", ""),
        ("Lote do Queijo", "lote_queijo", ""),
        ("N° Queijomatic", "numero_queijomatic", ""),
        ("Início Enchimento", "inicio_enchimento", ""),
        ("Quantidade Leite", "quantidade_leite", " L"),
        ("Temperatura. De Pasteurização", "temperatura_pasteurizacao", " °C"),
        ("Fosfatase", "fosfatase", "status"),
        ("Peroxidase", "peroxidase", "status"),
        ("Gordura Inicial", "gordura_inicial", " %"),
        ("Gordura Final", "gordura_final", " %"),
        ("Acidez", "acidez", " °D"),
        ("Temperatura da Coagulação", "temperatura_coagulacao", " °C"),
        ("Hora da Coagulação", "hora_coagulacao", ""),
        ("Hora do Corte", "hora_corte", ""),
        ("Temperatura de Cozimento", "temperatura_cozimento", " °C"),
        ("Responsável pela Produção", "responsavel", ""),
    ]

    for label, key, suffix in simple_fields:
        row = find_row_by_label(table, label)
        if row is None:
            continue
        row_cells = cells(row)
        for column, item in enumerate(items[:6], start=1):
            if column >= len(row_cells):
                continue
            if suffix == "status":
                cell_value = status_label(item.get(key))
            elif suffix:
                cell_value = value(item, key, suffix)
            else:
                cell_value = item.get(key) or ""
            set_cell_text(row_cells[column], cell_value)

    fermento_quantity_rows = find_rows_by_label(table, "Quantidade de Fermento")
    insumo_rows = {
        "fermento_mvd": (fermento_quantity_rows[0] if len(fermento_quantity_rows) > 0 else None, 0, "Lote do Fermento"),
        "fermento_fast": (fermento_quantity_rows[1] if len(fermento_quantity_rows) > 1 else None, 1, "Lote do Fermento"),
        "fermento": (fermento_quantity_rows[2] if len(fermento_quantity_rows) > 2 else None, 2, "Lote do Fermento"),
        "cloreto": (find_row_by_label(table, "Quantidade de Cloreto"), 0, "Lote do Cloreto"),
        "corante": (find_row_by_label(table, "Quantidade de Corante"), 0, "Lote do Corante"),
        "coalho": (find_row_by_label(table, "Quantidade de Coalho"), 0, "Lote do Coalho"),
    }

    for column, item in enumerate(items[:6], start=1):
        grouped = grouped_insumos(item)
        for kind, (quantity_row, lot_index, lot_label) in insumo_rows.items():
            current = grouped.get(kind, {})
            if quantity_row is not None:
                quantity_cells = cells(quantity_row)
                if column < len(quantity_cells):
                    set_cell_text(quantity_cells[column], "; ".join(current.get("quantidades", [])))
            lot_rows = find_rows_by_label(table, lot_label)
            if lot_index < len(lot_rows):
                lot_cells = cells(lot_rows[lot_index])
                if column < len(lot_cells):
                    set_cell_text(lot_cells[column], "; ".join(current.get("lotes", [])))


def fill_soro_refrigerado(root, data):
    table = find_table(root, 0)
    table_rows = rows(table)
    if len(table_rows) > 1:
        row_cells = cells(table_rows[1])
        values = [
            date_br(data.get("data_registro")),
            value(data, "entrada_diaria_estoque", " L"),
            value(data, "estoque_total", " L"),
            value(data, "litragem_vendida", " L"),
            value(data, "sobra_estoque", " L"),
            data.get("silo_armazenado") or "",
            data.get("responsavel") or "",
        ]
        for index, cell_value in enumerate(values):
            if index < len(row_cells):
                set_cell_text(row_cells[index], cell_value)

    if data.get("observacoes"):
        observations_table = find_table(root, 1)
        row_cells = cells(rows(observations_table)[0])
        if row_cells:
            append_to_label(row_cells[0], data.get("observacoes"))


def fill_formulacao_creme(root, data):
    header_table = find_table(root, 0)
    header_cells = cells(rows(header_table)[2])
    header_values = [
        data.get("responsavel_monitoramento") or "",
        month_label(data.get("mes")),
        value(data, "ano"),
        data.get("tipo_creme") or "",
    ]
    for index, cell_value in enumerate(header_values):
        if index < len(header_cells):
            append_to_label(header_cells[index], cell_value)

    data_table = find_table(root, 1)
    row_cells = cells(rows(data_table)[1])
    values = [
        date_br(data.get("data_fabricacao")),
        data.get("lote_creme_produzido") or "",
        value(data, "gordura_inicial", " %"),
        value(data, "gordura_final", " %"),
        value(data, "acidez", " °D"),
        data.get("responsavel") or "",
    ]
    for index, cell_value in enumerate(values):
        if index < len(row_cells):
            set_cell_text(row_cells[index], cell_value)


def fill_producao_creme(root, data):
    header_table = find_table(root, 0)
    header_cells = cells(rows(header_table)[2])
    header_values = [
        data.get("responsavel_monitoramento") or "",
        month_label(data.get("mes")),
        value(data, "ano"),
        data.get("tipo_creme") or "",
    ]
    for index, cell_value in enumerate(header_values):
        if index < len(header_cells):
            append_to_label(header_cells[index], cell_value)

    data_table = find_table(root, 1)
    row_cells = cells(rows(data_table)[1])
    values = [
        date_br(data.get("data_fabricacao")),
        data.get("lote_creme_produzido") or "",
        value(data, "quantidade_produzida_kg", " kg"),
        data.get("responsavel") or "",
    ]
    for index, cell_value in enumerate(values):
        if index < len(row_cells):
            set_cell_text(row_cells[index], cell_value)


FILLERS = {
    "formulacao_queijo": fill_formulacao_queijo,
    "soro_refrigerado": fill_soro_refrigerado,
    "formulacao_creme": fill_formulacao_creme,
    "producao_creme": fill_producao_creme,
}


def slug(value):
    text = re.sub(r"[^a-zA-Z0-9_-]+", "-", value or "").strip("-").lower()
    return text or "documento"


def output_name(tipo, data, formato):
    if tipo == "ordem_producao":
        key = data.get("data") or data.get("id") or "ordem"
        escopo = data.get("escopo") or "op"
        return f"ordem-producao-{slug(str(escopo))}-{slug(str(key))}.{formato}"

    config = FORM_CONFIG[tipo]
    if tipo == "formulacao_queijo" and isinstance(data.get("items"), list):
        key = data.get("data_formulacao") or data.get("id")
        return f"{config['documento'].lower().replace('.', '_')}_{config['slug']}_{slug(str(key))}.{formato}"
    key = data.get("lote_queijo") or data.get("lote_creme_produzido") or data.get("data_registro") or data.get("id")
    return f"{config['documento'].lower().replace('.', '_')}_{config['slug']}_{slug(str(key))}.{formato}"


def fill_docx(tipo, data, out_dir):
    config = FORM_CONFIG[tipo]
    template = TEMPLATE_DIR / config["template"]
    if not template.exists():
        raise FileNotFoundError(f"Modelo não encontrado: {template}")

    out_dir.mkdir(parents=True, exist_ok=True)
    docx_path = out_dir / output_name(tipo, data, "docx")

    with ZipFile(template, "r") as source:
        document_xml = source.read("word/document.xml")
        root = ET.fromstring(document_xml)
        FILLERS[tipo](root, data)
        updated_xml = ET.tostring(root, encoding="utf-8", xml_declaration=True)
        updated_xml = ensure_ignorable_namespace_declarations(updated_xml)
        logo_bytes = LOGO_PATH.read_bytes() if LOGO_PATH.exists() else None

        with ZipFile(docx_path, "w", ZIP_DEFLATED) as target:
            for item in source.infolist():
                if item.filename == LOGO_TARGET and logo_bytes:
                    continue
                content = source.read(item.filename)
                if item.filename == "word/document.xml":
                    content = updated_xml
                elif item.filename == "word/_rels/header1.xml.rels" and logo_bytes:
                    content = replace_header_logo_relationships(content)
                elif item.filename == "[Content_Types].xml" and logo_bytes:
                    content = ensure_png_content_type(content)
                target.writestr(item, content)

            if logo_bytes:
                target.writestr(LOGO_TARGET, logo_bytes)

    return docx_path


def replace_header_logo_relationships(content):
    namespace = "http://schemas.openxmlformats.org/package/2006/relationships"
    root = ET.fromstring(content)
    for rel in root.findall(f"{{{namespace}}}Relationship"):
        rel_type = rel.attrib.get("Type", "")
        target = rel.attrib.get("Target", "")
        if rel_type.endswith("/image") and target.startswith("media/"):
            rel.set("Target", "media/santilac_logo.png")
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def ensure_png_content_type(content):
    namespace = "http://schemas.openxmlformats.org/package/2006/content-types"
    root = ET.fromstring(content)
    has_png = any(
        child.tag == f"{{{namespace}}}Default" and child.attrib.get("Extension") == "png"
        for child in root
    )
    if not has_png:
        default = ET.Element(f"{{{namespace}}}Default")
        default.set("Extension", "png")
        default.set("ContentType", "image/png")
        root.insert(0, default)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def office_binary():
    return shutil.which("libreoffice") or shutil.which("soffice")


def convert_to_pdf(docx_path, out_dir):
    binary = office_binary()
    if binary is None:
        raise RuntimeError("LibreOffice/soffice não encontrado para gerar PDF.")

    with tempfile.TemporaryDirectory(prefix="santilac-lo-") as profile:
        command = [
            binary,
            "--headless",
            "--nologo",
            "--nofirststartwizard",
            f"-env:UserInstallation=file://{profile}",
            "--convert-to",
            "pdf",
            "--outdir",
            str(out_dir),
            str(docx_path),
        ]
        completed = subprocess.run(command, capture_output=True, text=True, timeout=90)

    if completed.returncode != 0:
        raise RuntimeError((completed.stderr or completed.stdout or "Falha ao converter PDF.").strip())

    pdf_path = out_dir / f"{docx_path.stem}.pdf"
    if not pdf_path.exists():
        generated = sorted(out_dir.glob("*.pdf"), key=lambda path: path.stat().st_mtime, reverse=True)
        if generated:
            generated[0].replace(pdf_path)

    if not pdf_path.exists():
        raise RuntimeError("PDF não foi criado pelo conversor.")

    docx_path.unlink(missing_ok=True)
    return pdf_path


def ordem_rows(ordem):
    values_by_label = {}
    duplicate_index = {}
    lote_queijo = lote_queijo_ordem(ordem)

    for campo in ordem.get("campos") or []:
        label = str(campo.get("rotulo") or "").strip()
        value_text = str(campo.get("valor") or "").strip()
        if not label:
            continue

        normalized = compact_normalize(label)
        current_index = duplicate_index.get(normalized, 0)
        duplicate_index[normalized] = current_index + 1
        values_by_label[(normalized, current_index)] = value_text

    seen = {}
    output = []
    for label in OP_FIELDS:
        normalized = compact_normalize(label)
        current_index = seen.get(normalized, 0)
        seen[normalized] = current_index + 1
        field_value = values_by_label.get((normalized, current_index), "")
        if normalized == compact_normalize("LOTE DO QUEIJO") and field_value == "":
            field_value = lote_queijo
        output.append((label, field_value))

    return output


def lote_queijo_ordem(ordem):
    lotes = []
    for formulacao in ordem.get("formulacoes") or []:
        lote = str(formulacao.get("lote_queijo") or "").strip()
        if lote and lote not in lotes:
            lotes.append(lote)

    return "; ".join(lotes)


def build_ordem_xlsx(data, out_dir):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    out_dir.mkdir(parents=True, exist_ok=True)
    output_path = out_dir / output_name("ordem_producao", data, "xlsx")
    ordens = data.get("ordens") or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Ordens de Produção"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    bold = Font(bold=True, name="Calibri", size=10)
    normal = Font(name="Calibri", size=10)

    for index, ordem in enumerate(ordens):
        start_col = 1 + (index * 3)
        label_col = start_col
        value_col = start_col + 1
        title_cell = ws.cell(row=1, column=label_col, value=f"OP {ordem.get('codigo_ordem') or ''}".strip())
        title_cell.font = Font(bold=True, name="Calibri", size=11)
        title_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=1, start_column=label_col, end_row=1, end_column=value_col)

        for row_index, (label, field_value) in enumerate(ordem_rows(ordem), start=2):
            left = ws.cell(row=row_index, column=label_col, value=label)
            right = ws.cell(row=row_index, column=value_col, value=field_value)
            for cell in (left, right):
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.font = normal
            left.font = bold
            if row_index == 2:
                left.fill = header_fill
                right.fill = header_fill

        ws.column_dimensions[get_column_letter(label_col)].width = 28
        ws.column_dimensions[get_column_letter(value_col)].width = 16
        ws.column_dimensions[get_column_letter(start_col + 2)].width = 3

    for row in range(1, len(OP_FIELDS) + 2):
        ws.row_dimensions[row].height = 20

    wb.save(output_path)
    return output_path


def build_ordem_pdf(data, out_dir):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table
    from reportlab.platypus.tables import TableStyle

    out_dir.mkdir(parents=True, exist_ok=True)
    output_path = out_dir / output_name("ordem_producao", data, "pdf")
    ordens = data.get("ordens") or []
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(A4),
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )

    label_style = ParagraphStyle("OpLabel", fontName="Helvetica-Bold", fontSize=6.8, leading=8)
    value_style = ParagraphStyle("OpValue", fontName="Helvetica-Bold", fontSize=7.2, leading=8.4)
    title_style = ParagraphStyle("OpTitle", fontName="Helvetica-Bold", fontSize=9, leading=10, alignment=1)

    story = []
    chunks = [ordens[index:index + 3] for index in range(0, len(ordens), 3)] or [[]]
    for chunk_index, chunk in enumerate(chunks):
        if chunk_index > 0:
            story.append(PageBreak())

        blocks = []
        for ordem in chunk:
            rows_pdf = [[Paragraph(f"OP {escape(ordem.get('codigo_ordem') or '')}", title_style), ""]]
            for label, field_value in ordem_rows(ordem):
                rows_pdf.append([
                    Paragraph(escape(label), label_style),
                    Paragraph(escape(field_value), value_style),
                ])
            table = Table(rows_pdf, colWidths=[47 * mm, 22 * mm])
            table.setStyle(TableStyle([
                ("SPAN", (0, 0), (1, 0)),
                ("BOX", (0, 0), (-1, -1), 0.65, colors.black),
                ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.black),
                ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#D9D9D9")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            blocks.append(table)

        while len(blocks) < 3:
            blocks.append("")

        page_table = Table([blocks], colWidths=[86 * mm, 86 * mm, 86 * mm])
        page_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.extend([page_table, Spacer(1, 2 * mm)])

    doc.build(story)
    return output_path


def run(args):
    with open(args.payload, "r", encoding="utf-8") as handle:
        data = json.load(handle)

    tipo = args.tipo
    formato = args.formato
    out_dir = Path(args.out_dir).resolve()

    if tipo == "ordem_producao":
        final_path = build_ordem_pdf(data, out_dir) if formato == "pdf" else build_ordem_xlsx(data, out_dir)
    else:
        final_path = build_pdf(tipo, data, out_dir) if formato == "pdf" else fill_docx(tipo, data, out_dir)

    return {
        "success": True,
        "arquivo": final_path.name,
        "caminho": str(final_path),
        "formato": formato,
        "errors": [],
    }


def build_pdf(tipo, data, out_dir):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table
    from reportlab.platypus.tables import TableStyle

    config = FORM_CONFIG[tipo]
    out_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = out_dir / output_name(tipo, data, "pdf")
    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    normal = ParagraphStyle("NormalSantiLac", parent=styles["Normal"], fontName="Helvetica", fontSize=8, leading=10)
    title = ParagraphStyle("TitleSantiLac", parent=styles["Heading1"], fontName="Helvetica-Bold", fontSize=13, leading=15, alignment=1)
    section_title = ParagraphStyle("SectionSantiLac", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=9, leading=11, textColor=colors.HexColor("#111111"))

    story = []
    logo = Image(str(LOGO_PATH), width=42 * mm, height=16 * mm) if LOGO_PATH.exists() else Paragraph("SantiLac", title)
    header = Table(
        [[logo, Paragraph(title_for_pdf(tipo), title), Paragraph(f"{config['documento']}<br/>Registro {data.get('id', '')}", normal)]],
        colWidths=[52 * mm, 160 * mm, 48 * mm],
    )
    header.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.7, colors.black),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, 0), "CENTER"),
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.extend([header, Spacer(1, 5 * mm)])

    if tipo == "formulacao_queijo" and isinstance(data.get("items"), list):
        labels = [
            ("Tipo de Queijo", "tipo_queijo", ""),
            ("Silo", "silo", ""),
            ("Lote do Leite", "lote_leite", ""),
            ("Lote do Queijo", "lote_queijo", ""),
            ("N° Queijomatic", "numero_queijomatic", ""),
            ("Início Enchimento", "inicio_enchimento", ""),
            ("Quantidade Leite", "quantidade_leite", " L"),
            ("Temperatura. De Pasteurização", "temperatura_pasteurizacao", " °C"),
            ("Fosfatase", "fosfatase", "status"),
            ("Peroxidase", "peroxidase", "status"),
            ("Gordura Inicial", "gordura_inicial", " %"),
            ("Gordura Final", "gordura_final", " %"),
            ("Acidez", "acidez", " °D"),
            ("Temperatura da Coagulação", "temperatura_coagulacao", " °C"),
            ("Hora da Coagulação", "hora_coagulacao", ""),
            ("Hora do Corte", "hora_corte", ""),
            ("Temperatura de Cozimento", "temperatura_cozimento", " °C"),
            ("Responsável pela Produção", "responsavel", ""),
        ]
        items = data.get("items") or []
        chunks = [items[index:index + 6] for index in range(0, len(items), 6)] or [[]]
        small = ParagraphStyle("SmallSantiLac", parent=normal, fontName="Helvetica", fontSize=6.2, leading=7.1)
        for chunk_index, chunk in enumerate(chunks):
            if chunk_index > 0:
                story.extend([PageBreak(), header, Spacer(1, 5 * mm)])
            rows_pdf = []
            for label, key, suffix in labels:
                left = f"{label}<br/>Data: {date_br(data.get('data_formulacao'))}" if label == "Tipo de Queijo" else label
                current_row = [Paragraph(f"<b>{escape(left)}</b>", small)]
                for item in chunk:
                    if suffix == "status":
                        current_row.append(Paragraph(escape(status_label(item.get(key)) or "-"), small))
                    elif suffix:
                        current_row.append(Paragraph(escape(value(item, key, suffix) or "-"), small))
                    else:
                        current_row.append(Paragraph(escape(item.get(key) or "-"), small))
                while len(current_row) < 7:
                    current_row.append(Paragraph("", small))
                rows_pdf.append(current_row)
            table = Table(rows_pdf, colWidths=[48 * mm] + [34 * mm] * 6)
            table.setStyle(form_table_style(colors))
            story.append(table)
        doc.build(story)
        return pdf_path

    for section, fields in pdf_sections(tipo, data):
        story.append(Paragraph(section, section_title))
        table_data = []
        row = []
        for label, field_value in fields:
            row.append(Paragraph(f"<b>{escape(label)}</b><br/>{escape(field_value or '-')}", normal))
            if len(row) == 4:
                table_data.append(row)
                row = []
        if row:
            while len(row) < 4:
                row.append(Paragraph("", normal))
            table_data.append(row)
        table = Table(table_data, colWidths=[65 * mm] * 4)
        table.setStyle(form_table_style(colors))
        story.extend([table, Spacer(1, 4 * mm)])

    if tipo == "formulacao_queijo":
        story.append(Paragraph("Insumos", section_title))
        insumos = [["Tipo", "Quantidade", "Unidade", "Lote"]]
        for item in data.get("insumos") or []:
            insumos.append([
                insumo_label(item.get("tipo_insumo")),
                value({"v": item.get("quantidade")}, "v"),
                item.get("unidade") or "-",
                item.get("lote_insumo") or "-",
            ])
        if len(insumos) == 1:
            insumos.append(["-", "-", "-", "-"])
        insumos_table = Table(insumos, colWidths=[65 * mm] * 4)
        insumos_table.setStyle(form_table_style(colors, header=True))
        story.extend([insumos_table, Spacer(1, 4 * mm)])

    observations = data.get("observacoes") or "Sem observações."
    story.append(Paragraph("Observações", section_title))
    obs_table = Table([[Paragraph(escape(observations), normal)]], colWidths=[260 * mm])
    obs_table.setStyle(form_table_style(colors))
    story.append(obs_table)

    doc.build(story)
    return pdf_path


def title_for_pdf(tipo):
    return {
        "formulacao_queijo": "Controle de Formulação do Queijo",
        "soro_refrigerado": "Controle de Produção de Soro Refrigerado",
        "formulacao_creme": "Controle de Formulação Creme",
        "producao_creme": "Controle de Produção Creme",
    }[tipo]


def pdf_sections(tipo, data):
    if tipo == "formulacao_queijo":
        return [
            ("Identificação", [
                ("Tipo de Queijo", data.get("tipo_queijo")),
                ("Data", date_br(data.get("data_formulacao"))),
                ("Silo", data.get("silo")),
                ("Lote do Leite", data.get("lote_leite")),
                ("Lote do Queijo", data.get("lote_queijo")),
                ("Nº Queijomatic", data.get("numero_queijomatic")),
                ("Início Enchimento", data.get("inicio_enchimento")),
                ("Quantidade Leite", value(data, "quantidade_leite", " L")),
                ("Temperatura de Pasteurização", value(data, "temperatura_pasteurizacao", " °C")),
                ("Fosfatase", status_label(data.get("fosfatase"))),
                ("Peroxidase", status_label(data.get("peroxidase"))),
            ]),
            ("Parâmetros de Fabricação", [
                ("Gordura Inicial", value(data, "gordura_inicial", " %")),
                ("Gordura Final", value(data, "gordura_final", " %")),
                ("Acidez", value(data, "acidez", " °D")),
                ("Temperatura da Coagulação", value(data, "temperatura_coagulacao", " °C")),
                ("Hora da Coagulação", data.get("hora_coagulacao")),
                ("Hora do Corte", data.get("hora_corte")),
                ("Temperatura de Cozimento", value(data, "temperatura_cozimento", " °C")),
            ]),
        ]
    if tipo == "soro_refrigerado":
        return [("Registro", [
            ("Data", date_br(data.get("data_registro"))),
            ("Entrada Diária no Estoque", value(data, "entrada_diaria_estoque", " L")),
            ("Estoque Total", value(data, "estoque_total", " L")),
            ("Litragem Vendida", value(data, "litragem_vendida", " L")),
            ("Sobra no Estoque", value(data, "sobra_estoque", " L")),
            ("Silo Armazenado", data.get("silo_armazenado")),
            ("Responsável", data.get("responsavel")),
        ])]
    if tipo == "formulacao_creme":
        return [("Registro", [
            ("Responsável pelo Monitoramento", data.get("responsavel_monitoramento")),
            ("Mês", month_label(data.get("mes"))),
            ("Ano", value(data, "ano")),
            ("Tipo de Creme", data.get("tipo_creme")),
            ("Fabricação", date_br(data.get("data_fabricacao"))),
            ("Lote do Creme Produzido", data.get("lote_creme_produzido")),
            ("Gordura Inicial", value(data, "gordura_inicial", " %")),
            ("Gordura Final", value(data, "gordura_final", " %")),
            ("Acidez", value(data, "acidez", " °D")),
            ("Responsável", data.get("responsavel")),
        ])]
    return [("Registro", [
        ("Responsável pelo Monitoramento", data.get("responsavel_monitoramento")),
        ("Mês", month_label(data.get("mes"))),
        ("Ano", value(data, "ano")),
        ("Tipo de Creme", data.get("tipo_creme")),
        ("Fabricação", date_br(data.get("data_fabricacao"))),
        ("Lote do Creme Produzido", data.get("lote_creme_produzido")),
        ("Quantidade Produzida", value(data, "quantidade_produzida_kg", " kg")),
        ("Responsável", data.get("responsavel")),
    ])]


def form_table_style(colors, header=False):
    from reportlab.platypus.tables import TableStyle

    commands = [
        ("BOX", (0, 0), (-1, -1), 0.6, colors.black),
        ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#555555")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]
    if header:
        commands.extend([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eeeeee")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ])
    return TableStyle(commands)


def insumo_label(kind):
    return {
        "fermento_mvd": "Fermento MVD",
        "fermento_fast": "Fermento FAST",
        "fermento": "Fermento",
        "cloreto": "Cloreto",
        "corante": "Corante",
        "coalho": "Coalho",
        "outro": "Outro",
    }.get(kind or "outro", kind or "Outro")


def escape(raw):
    return str(raw or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def main():
    parser = argparse.ArgumentParser(description="Processador de formulários de produção SantiLac.")
    parser.add_argument("--tipo", required=True, choices=sorted([*FORM_CONFIG.keys(), "ordem_producao"]))
    parser.add_argument("--payload", required=True)
    parser.add_argument("--out-dir", required=True)
    parser.add_argument("--format", dest="formato", required=True, choices=["docx", "pdf", "xlsx"])
    args = parser.parse_args()

    try:
        result = run(args)
    except Exception as exc:
        result = {
            "success": False,
            "arquivo": None,
            "caminho": None,
            "formato": args.formato,
            "errors": [str(exc)],
        }

    print(json.dumps(result, ensure_ascii=False))
    return 0 if result["success"] else 1


if __name__ == "__main__":
    sys.exit(main())
